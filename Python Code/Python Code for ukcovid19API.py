# -*- coding: utf-8 -*-
"""
Created on Fri May 14 23:38:53 2021

@author: Karthik
"""

#---Import required libraries
from uk_covid19 import Cov19API
import datetime
import pandas as pd
import os
import openpyxl as xl


#---Create an empty dataframe
dtf = pd.DataFrame()


#---Create the list of parameters for data request
parameters = {
   	"areaType":"areaType",
	"areaName":"areaName",
	"areaCode":"areaCode",
	"date":"date",
	"newCasesByPublishDate":"newCasesByPublishDate",
	"cumCasesByPublishDate":"cumCasesByPublishDate",
	"cumCasesByPublishDateRate":"cumCasesByPublishDateRate",
	"newCasesBySpecimenDate":"newCasesBySpecimenDate",
	"cumCasesBySpecimenDate":"cumCasesBySpecimenDate",
	"cumCasesBySpecimenDateRate":"cumCasesBySpecimenDateRate",
	"maleCases":"maleCases",
	"femaleCases":"femaleCases",
	"newAdmissions":"newAdmissions",
	"cumAdmissions":"cumAdmissions",
	"cumAdmissionsByAge":"cumAdmissionsByAge",
	"cumTestsByPublishDate":"cumTestsByPublishDate",
	"newTestsByPublishDate":"newTestsByPublishDate",
	"covidOccupiedMVBeds":"covidOccupiedMVBeds",
	"hospitalCases":"hospitalCases",
	"newDeaths28DaysByPublishDate":"newDeaths28DaysByPublishDate",
	"cumDeaths28DaysByPublishDate":"cumDeaths28DaysByPublishDate",
	"cumDeaths28DaysByPublishDateRate":"cumDeaths28DaysByPublishDateRate",
	"newDeaths28DaysByDeathDate":"newDeaths28DaysByDeathDate",
	"cumDeaths28DaysByDeathDate":"cumDeaths28DaysByDeathDate",
	"cumDeaths28DaysByDeathDateRate":"cumDeaths28DaysByDeathDateRate"
    }


#---Request and get data for past 7 days
for i in range(1,8):
    req_day = (datetime.datetime.now() - datetime.timedelta(days=(i-1))).strftime("%Y-%m-%d")
    query_filters = ['areaType=ltla', f'date={req_day}' ]

    if dtf.empty:
        dtf = Cov19API(filters=query_filters, structure=parameters).get_dataframe()
    else:
        temp_df = Cov19API(filters=query_filters, structure=parameters).get_dataframe()
        dtf = dtf.append(temp_df, ignore_index=True)

dtf.reset_index(drop=True, inplace=True)
print(dtf.head())

        
#---Save to a new file if file does not exist
if os.path.isfile(r"C:\Users\Karthik\Desktop\UKCovid19 API Codes\Python Excel File\UK_7day_Covid_Data.xlsx") == False:
    writer = pd.ExcelWriter(r'C:\Users\Karthik\Desktop\UKCovid19 API Codes\Python Excel File\UK_7day_Covid_Data.xlsx', engine='openpyxl')
    dtf.to_excel(writer, sheet_name='UK_7day_Covid_API_Data', index=False)
    writer.save()


#---If files exists, clear existing data and update with new data
else: 
    wb = xl.load_workbook(r"C:\Users\Karthik\Desktop\UKCovid19 API Codes\Python Excel File\UK_7day_Covid_Data.xlsx")
    if 'UK_7day_Covid_API_Data' in wb.sheetnames:
        ws = wb['UK_7day_Covid_API_Data']
        ws.delete_rows(1, ws.max_row)
        wb.save(r"C:\Users\Karthik\Desktop\UKCovid19 API Codes\Python Excel File\UK_7day_Covid_Data.xlsx")
 
    with pd.ExcelWriter(r'C:\Users\Karthik\Desktop\UKCovid19 API Codes\Python Excel File\UK_7day_Covid_Data.xlsx', engine="openpyxl", mode='a') as writer:  
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        dtf.to_excel(writer, sheet_name='UK_7day_Covid_API_Data', startrow=0, startcol=0, index=False)
        writer.save()

    
        
    
    